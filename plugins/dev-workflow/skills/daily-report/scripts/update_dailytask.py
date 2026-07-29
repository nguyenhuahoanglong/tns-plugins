"""Update the newest-first daily-report workbook without relying on a caller cwd."""
import argparse
import copy
import datetime
import os
import time
import zipfile
from pathlib import Path

import openpyxl

from lib_common import load_config


COL = {"date": "A", "yesterday": "B", "today": "C", "report": "E", "timesheet": "F"}


def build_report(yesterday, today):
    lines = ["Yesterday"]
    if str(yesterday or "").strip():
        lines.append(str(yesterday).strip())
    lines.append("Today")
    if str(today or "").strip():
        lines.append(str(today).strip())
    return "\n".join(lines)


def build_timesheet_memo(cfg, today_text):
    defaults = cfg["timesheet"]["defaults"]
    days = defaults["task_days"]
    days_text = str(int(days)) if float(days).is_integer() else str(days)
    code = defaults["bindings"]["cr90e_ProjectCodeCD"]["code"]
    return (f"Task Date: Today\nProject Code: {code}\nLocation: {defaults['location_label']}\n"
            f"Task Days: {days_text}\nTask Description:\n{today_text}")


def _copy_style(source, destination):
    destination.font = copy.copy(source.font)
    destination.alignment = copy.copy(source.alignment)
    destination.fill = copy.copy(source.fill)
    destination.border = copy.copy(source.border)
    destination.number_format = source.number_format


# Excel's day 1 is 1900-01-01 and it treats 1900 as a leap year, so serials count from
# 1899-12-30. That is exact for any date after 1900-02-28, which every report date is.
_EXCEL_EPOCH = datetime.datetime(1899, 12, 30)


def _as_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # A date cell that lost its number format reads back as a bare serial. Decoding
        # it lets a workbook already damaged that way match its own row again instead of
        # inserting a second row for the same day.
        try:
            return (_EXCEL_EPOCH + datetime.timedelta(days=float(value))).date()
        except (OverflowError, ValueError):
            return None
    return None


def _date_number_format(sheet, column, prior_row):
    """Pick a date format that survives a save/load round trip.

    A same-day refresh in a workbook with one data row has no row beneath it, and
    reading a missing cell yields ``General``. Adopting that would serialise the date as
    a bare number, so the following run would not recognise the row and would insert a
    duplicate for the same day. The header carries the intended date format.
    """
    candidates = []
    if prior_row <= sheet.max_row:
        candidates.append(sheet[f"{column}{prior_row}"].number_format)
    candidates.append(sheet[f"{column}1"].number_format)
    for candidate in candidates:
        if candidate and candidate != "General":
            return candidate
    return "yyyy-mm-dd"


def _unique_lines(*blocks):
    """Merge non-empty lines in first-seen order for same-day idempotence."""
    result = []
    seen = set()
    for block in blocks:
        if not block:
            continue
        for line in str(block).splitlines():
            line = line.strip()
            if line and line not in seen:
                seen.add(line)
                result.append(line)
    return "\n".join(result)


def _temp_path(path):
    return path.with_name(f".{path.stem}.tmp{path.suffix}")


def _remove_quietly(path):
    try:
        Path(path).unlink()
    except FileNotFoundError:
        pass
    except OSError:
        pass


def _persist_explicit_empty_yesterday(path, sheet_name):
    """Keep an intentional first-row empty string distinct from an absent cell.

    openpyxl omits empty strings when serialising; its reader does preserve an
    explicit inline string, which is useful to consumers of the report contract.
    """
    path = Path(path)
    if not path.exists():
        return
    rewritten = path.with_name(path.name + ".rewrite")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(rewritten, "w", zipfile.ZIP_DEFLATED) as destination:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename.startswith("xl/worksheets/"):
                text = data.decode("utf-8")
                marker = 'r="B2" t="inlineStr"></c>'
                if marker in text:
                    text = text.replace(marker, 'r="B2" t="inlineStr"><is><t></t></is></c>')
                    data = text.encode("utf-8")
            destination.writestr(item, data)
    os.replace(rewritten, path)


def update_workbook(config, today, today_additions=(), yesterday_addition="", date=None,
                    workbook_loader=openpyxl.load_workbook, save_workbook=None,
                    sleep=time.sleep, replace_file=os.replace):
    """Persist one portable workbook update and return a structured result.

    The configured ``excel.path`` is deliberately the only workbook location used.
    Saving goes to a sibling temporary file before an atomic replacement, so a locked
    or failed save cannot damage the prior workbook.
    """
    path = Path(config["excel"]["path"]).expanduser()
    sheet_name = config["excel"]["sheet"]
    report_date = date or datetime.date.today()
    if isinstance(report_date, datetime.datetime):
        report_date = report_date.date()
    target_date = datetime.datetime(report_date.year, report_date.month, report_date.day)
    save_workbook = save_workbook or (lambda workbook, target: workbook.save(target))
    retry_delay_seconds = 1

    try:
        workbook = workbook_loader(path)
        sheet = workbook[sheet_name]
    except Exception as error:
        return {"status": "FAIL", "code": "WORKBOOK_LOAD_FAILED", "message": str(error)}

    had_data_row = sheet.max_row >= 2 and any(sheet.cell(2, column).value is not None for column in range(1, 7))
    existing_today = _as_date(sheet["A2"].value) == report_date if had_data_row else False
    if existing_today:
        row, prior_row = 2, 3
        prior_today = sheet[f"{COL['today']}{row}"].value or ""
    else:
        row, prior_row = 2, 3
        sheet.insert_rows(row)
        if had_data_row:
            for column in COL.values():
                _copy_style(sheet[f"{column}{prior_row}"], sheet[f"{column}{row}"])
        else:
            # Task 9 creates headers only.  Header A1 supplies its date format for
            # the first row, while the other cells keep normal workbook defaults.
            sheet[f"{COL['date']}{row}"].number_format = sheet[f"{COL['date']}1"].number_format
        prior_today = ""

    yesterday = "" if existing_today and prior_row > sheet.max_row else (sheet[f"{COL['today']}{prior_row}"].value or "")
    if yesterday_addition.strip():
        yesterday = _unique_lines(yesterday, yesterday_addition)
    today_text = _unique_lines(prior_today if existing_today else today, today if existing_today else "", *today_additions)
    # For a new day, today is the authoritative base; for a same-day refresh it is
    # retained and additions are merged without duplicating prompt lines.
    if not existing_today:
        today_text = _unique_lines(today, *today_additions)
    report = build_report(yesterday, today_text)
    memo = build_timesheet_memo(config, today_text)

    sheet[f"{COL['date']}{row}"] = target_date
    if had_data_row:
        sheet[f"{COL['date']}{row}"].number_format = _date_number_format(sheet, COL["date"], prior_row)
    sheet[f"{COL['yesterday']}{row}"] = yesterday
    sheet[f"{COL['today']}{row}"] = today_text
    sheet[f"{COL['report']}{row}"] = report
    sheet[f"{COL['timesheet']}{row}"] = memo

    temporary = _temp_path(path)
    for attempt in range(2):
        try:
            save_workbook(workbook, temporary)
            if yesterday == "":
                _persist_explicit_empty_yesterday(temporary, sheet_name)
            replace_file(temporary, path)
            return {"status": "UPDATED", "path": str(path), "mode": "UPDATE" if existing_today else "INSERT",
                    "report": report, "today": today_text, "retry_delay_seconds": retry_delay_seconds}
        except PermissionError:
            _remove_quietly(temporary)
            if attempt == 0:
                sleep(retry_delay_seconds)
                continue
            return {"status": "FAIL", "code": "WORKBOOK_LOCKED",
                    "message": "Workbook is locked. Close it and try again.", "retry_delay_seconds": retry_delay_seconds}
        except OSError as error:
            return {"status": "FAIL", "code": "WORKBOOK_SAVE_FAILED",
                    "message": f"Workbook temporary save or atomic replacement failed: {error}"}
        finally:
            _remove_quietly(temporary)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--today", required=True)
    parser.add_argument("--yesterday-add", default="")
    parser.add_argument("--date")
    parser.add_argument("--config")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    config = load_config(args.config)
    date = datetime.datetime.strptime(args.date, "%Y-%m-%d").date() if args.date else None
    if args.dry_run:
        # Preserve the legacy no-write CLI boundary while sharing formatting semantics.
        workbook = openpyxl.load_workbook(config["excel"]["path"])
        sheet = workbook[config["excel"]["sheet"]]
        yesterday = sheet["C2"].value or "" if sheet.max_row >= 2 else ""
        print("=== WOULD WRITE (row 2) ===")
        print(build_report(yesterday, args.today.strip("\n")))
        return
    result = update_workbook(config, args.today, yesterday_addition=args.yesterday_add, date=date)
    if result["status"] != "UPDATED":
        raise SystemExit(f"{result['code']}: {result['message']}")
    print("=== WRITING (row 2) ===")
    print(f"Saved: {result['path']}")
    print("\n=== TEAMS REPORT (copy below) ===")
    print(result["report"])


if __name__ == "__main__":
    main()
