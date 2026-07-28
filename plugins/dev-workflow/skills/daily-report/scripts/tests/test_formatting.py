#!/usr/bin/env python3
"""Dependency-free unit tests for the daily-report pure formatting logic.

Runs under pytest (`python -m pytest scripts/tests/`) and standalone.
The live-system behaviour (ADO query, Excel write, Dataverse) is verified by
verify_output.py against real outputs, not here.
"""
# Test registry: .plans/portable-git-daily-report-dev-workflow.daily.test-cases.md
# Subject: portable workbook update and optional workbook backup (Task 11)
import datetime
import inspect
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).resolve().parent.parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import update_dailytask as ud  # noqa: E402
import write_timesheet as wt  # noqa: E402
import commit_workbook as cw  # noqa: E402
import pending_timesheets as pt  # noqa: E402

CFG = {
    "timesheet": {
        "defaults": {
            "task_days": 1.0,
            "location_label": "Office",
            "from_hour_local": "09:00",
            "to_hour_local": "18:00",
            "timezone_offset_hours": 7,
            "bindings": {"cr90e_ProjectCodeCD": {"code": "RDDMS01"}},
        }
    }
}


def _portable_cfg(workbook):
    """Return only portable workbook settings plus the established memo defaults."""
    return {
        "excel": {"path": str(workbook), "sheet": "Daily Report"},
        **CFG,
    }


def _headers_only_workbook(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Daily Report"
    sheet.append(["Date", "Yesterday", "Today", "", "Report", "Timesheet"])
    sheet["A1"].font = openpyxl.styles.Font(bold=True)
    sheet["A1"].number_format = "yyyy-mm-dd"
    workbook.save(path)


def _require_parameters(function, *names):
    parameters = inspect.signature(function).parameters
    missing = [name for name in names if name not in parameters]
    assert not missing, f"Task 11 must expose portable injection seams: {', '.join(missing)}."


def test_report_template():
    r = ud.build_report("- a", "- b\n- c")
    assert r == "Yesterday\n- a\nToday\n- b\n- c"


def test_timesheet_memo_matches_existing_pattern():
    memo = ud.build_timesheet_memo(CFG, "- #1 task")
    assert memo == ("Task Date: Today\nProject Code: RDDMS01\nLocation: Office\n"
                    "Task Days: 1\nTask Description:\n- #1 task")


def test_description_transform_matches_portal_pattern():
    # multi-item: strip "- ", join with "; " (matches the real 2026-06-09 portal entry)
    block = "- #414981 Service Plan Deal: Design quote carry-forward\n- Sprint Planning"
    assert wt.format_description(block) == "#414981 Service Plan Deal: Design quote carry-forward; Sprint Planning"
    # single item
    assert wt.format_description("- #2168 MDM Upload") == "#2168 MDM Upload"
    # verbatim style keeps bullets
    assert wt.format_description(block, "verbatim") == block


def test_local_hour_to_utc_applies_offset():
    d = datetime.date(2026, 6, 10)
    assert wt.local_hour_to_utc(d, "09:00", 7) == "2026-06-10T02:00:00Z"
    assert wt.local_hour_to_utc(d, "18:00", 7) == "2026-06-10T11:00:00Z"


class _FakeDv:
    """Minimal Dataverse stand-in: returns canned rows for resolve_header's GET."""
    def __init__(self, rows):
        self._rows = rows

    def get(self, path, prefer=None):
        return {"json": {"value": self._rows}}


_HDR_CFG = {"timesheet": {"employee_id": "emp-1", "header_entity_set": "headers"}}


def test_resolve_header_single_match_returns_row():
    row = {"cr90e_refnbr": "TS-01", "cr90e_xts_timesheet_timesheetheaderid": "h1"}
    assert wt.resolve_header(_FakeDv([row]), _HDR_CFG, datetime.date(2026, 6, 15)) is row


def test_resolve_header_no_match_raises():
    try:
        wt.resolve_header(_FakeDv([]), _HDR_CFG, datetime.date(2026, 6, 15))
        assert False, "expected SystemExit"
    except SystemExit as e:
        assert "no active timesheet period header" in str(e)


def test_resolve_header_ambiguous_lists_candidates():
    rows = [
        {"cr90e_refnbr": "TS-01", "cr90e_fromperiod": "2026-06-01", "cr90e_toperiod": "2026-06-30"},
        {"cr90e_refnbr": "TS-02", "cr90e_fromperiod": "2026-06-10", "cr90e_toperiod": "2026-06-20"},
    ]
    try:
        wt.resolve_header(_FakeDv(rows), _HDR_CFG, datetime.date(2026, 6, 15))
        assert False, "expected SystemExit"
    except SystemExit as e:
        msg = str(e)
        assert "AMBIGUOUS_PERIOD" in msg
        assert "TS-01" in msg and "TS-02" in msg  # both candidates surfaced


def _git(repo, *args):
    return subprocess.run(
        ["git", "-C", str(repo), *args],
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()


def test_commit_workbook_commits_only_workbook():
    with tempfile.TemporaryDirectory() as temp_dir:
        repo = Path(temp_dir)
        _git(repo, "init")
        _git(repo, "config", "user.name", "Daily Report Test")
        _git(repo, "config", "user.email", "daily-report@example.test")

        workbook = repo / "DailyTask.xlsx"
        unrelated = repo / "notes.md"
        workbook.write_bytes(b"version 1")
        unrelated.write_text("version 1", encoding="utf-8")
        _git(repo, "add", ".")
        _git(repo, "commit", "-m", "initial")

        workbook.write_bytes(b"version 2")
        unrelated.write_text("version 2", encoding="utf-8")
        _git(repo, "add", "notes.md")

        commit_hash = cw.commit_workbook(workbook, "chore: update daily task workbook")

        assert commit_hash
        assert _git(repo, "show", "--pretty=", "--name-only", "HEAD") == "DailyTask.xlsx"
        assert _git(repo, "diff", "--cached", "--name-only") == "notes.md"
        assert _git(repo, "status", "--short") == "M  notes.md"


def test_pending_enqueue_creates_and_updates_one_date():
    with tempfile.TemporaryDirectory() as temp_dir:
        queue = Path(temp_dir) / "pending-timesheets.json"

        rec, action = pt.enqueue(queue, "2026-07-01", "- #1 First", "missing period")
        assert action == "CREATED"
        assert rec["status"] == "pending"
        assert rec["description"] == "#1 First"

        rec, action = pt.enqueue(queue, "2026-07-01", "- #2 Second", "still missing")
        data = pt.load_queue(queue)
        assert action == "UPDATED"
        assert len(data["records"]) == 1
        assert data["records"][0]["todayBlock"] == "- #2 Second"
        assert data["records"][0]["lastError"] == "still missing"


def test_pending_sync_marks_successful_record_synced():
    with tempfile.TemporaryDirectory() as temp_dir:
        queue = Path(temp_dir) / "pending-timesheets.json"
        pt.enqueue(queue, "2026-07-01", "- #1 First", "missing period")
        calls = []

        def runner(cmd, cwd):
            calls.append(cmd)
            return subprocess.CompletedProcess(cmd, 0, stdout="OK", stderr="")

        result = pt.sync(queue, runner=runner, script_path=SCRIPTS_DIR / "write_timesheet.py")
        data = pt.load_queue(queue)

        assert result["synced"] == 1
        assert data["records"][0]["status"] == "synced"
        assert data["records"][0]["syncedAt"]
        assert any("--check-auth" in c for c in calls)
        assert any("--commit" in c for c in calls)


def test_pending_sync_keeps_missing_period_pending():
    with tempfile.TemporaryDirectory() as temp_dir:
        queue = Path(temp_dir) / "pending-timesheets.json"
        pt.enqueue(queue, "2026-07-01", "- #1 First", "missing period")

        def runner(cmd, cwd):
            if "--check-auth" in cmd:
                return subprocess.CompletedProcess(cmd, 0, stdout="AUTH_OK", stderr="")
            return subprocess.CompletedProcess(
                cmd, 1,
                stdout="",
                stderr="ERROR: no active timesheet period header found for 2026-07-01.",
            )

        result = pt.sync(queue, runner=runner, script_path=SCRIPTS_DIR / "write_timesheet.py")
        data = pt.load_queue(queue)

        assert result["failed"] == 1
        assert data["records"][0]["status"] == "pending"
        assert data["records"][0]["attempts"] == 1
        assert "no active timesheet period" in data["records"][0]["lastError"]


def test_pending_prune_removes_old_synced_records_only():
    with tempfile.TemporaryDirectory() as temp_dir:
        queue = Path(temp_dir) / "pending-timesheets.json"
        old = (datetime.datetime.now() - datetime.timedelta(days=31)).isoformat(timespec="seconds")
        recent = datetime.datetime.now().isoformat(timespec="seconds")
        pt.save_queue(queue, {"version": 1, "records": [
            {"date": "2026-06-01", "status": "synced", "syncedAt": old},
            {"date": "2026-07-01", "status": "synced", "syncedAt": recent},
            {"date": "2026-07-02", "status": "pending", "syncedAt": None},
        ]})

        removed = pt.prune(queue, 30)
        data = pt.load_queue(queue)

        assert [r["date"] for r in removed] == ["2026-06-01"]
        assert [r["date"] for r in data["records"]] == ["2026-07-01", "2026-07-02"]


# TC-053: Create the first report row from a portable headers-only workbook.
# Steps:
#   1. Create a workbook with only the Task 9 generated headers.
#   2. Update it through the portable config workbook path.
#   3. Verify the first newest-first row, empty Yesterday, columns, formulas, and formatting.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-5, AC-9, AC-12.
def test_tc_053_creates_first_report_row_from_headers_only_workbook(tmp_path):
    workbook = tmp_path / "DailyTask.xlsx"
    _headers_only_workbook(workbook)
    cfg = _portable_cfg(workbook)

    _require_parameters(ud.update_workbook, "config", "today", "date", "workbook_loader", "save_workbook")
    result = ud.update_workbook(
        config=cfg, today="- #101 First task", date=datetime.date(2026, 7, 27),
        workbook_loader=openpyxl.load_workbook, save_workbook=lambda wb, path: wb.save(path),
    )

    saved = openpyxl.load_workbook(workbook, data_only=False)
    sheet = saved["Daily Report"]
    assert result["status"] == "UPDATED"
    assert sheet.max_row == 2
    assert [sheet[f"{column}1"].value for column in ("A", "B", "C", "E", "F")] == [
        "Date", "Yesterday", "Today", "Report", "Timesheet"
    ]
    assert sheet["A2"].value.date() == datetime.date(2026, 7, 27)
    assert sheet["A2"].number_format == "yyyy-mm-dd"
    assert sheet["B2"].value == ""
    assert sheet["C2"].value == "- #101 First task"
    assert sheet["E2"].value == "Yesterday\n\nToday\n- #101 First task"
    assert sheet["F2"].value == ud.build_timesheet_memo(cfg, "- #101 First task")


# TC-054: Repeating a date updates the same row and merges prompt additions deterministically.
# Steps:
#   1. Update a temporary workbook twice for one date with ordered prompt additions.
#   2. Read the resulting workbook.
#   3. Verify one row and one stable, de-duplicated Today block.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-9, AC-12.
def test_tc_054_updates_same_date_without_duplicate_row_and_merges_prompt_additions(tmp_path):
    workbook = tmp_path / "DailyTask.xlsx"
    _headers_only_workbook(workbook)
    cfg = _portable_cfg(workbook)
    _require_parameters(ud.update_workbook, "config", "today", "today_additions", "date", "workbook_loader", "save_workbook")

    for additions in (("- Planning",), ("- Planning", "- Review")):
        ud.update_workbook(
            config=cfg, today="- #101 First task", today_additions=additions,
            date=datetime.date(2026, 7, 27), workbook_loader=openpyxl.load_workbook,
            save_workbook=lambda wb, path: wb.save(path),
        )

    sheet = openpyxl.load_workbook(workbook)["Daily Report"]
    assert sheet.max_row == 2
    assert sheet["C2"].value == "- #101 First task\n- Planning\n- Review"
    assert sheet["E2"].value.endswith("Today\n- #101 First task\n- Planning\n- Review")


# TC-055: A locked save retries once through injected save/sleep boundaries and then succeeds.
# Steps:
#   1. Inject a first save failure and a successful second save.
#   2. Update a temporary workbook.
#   3. Verify exactly one retry and no real wait.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-9, AC-12.
def test_tc_055_retries_locked_workbook_once_then_succeeds_without_real_sleep(tmp_path):
    workbook = tmp_path / "DailyTask.xlsx"
    _headers_only_workbook(workbook)
    saves, sleeps = [], []

    def save_once_locked(wb, path):
        saves.append(Path(path))
        if len(saves) == 1:
            raise PermissionError("locked")
        wb.save(path)

    _require_parameters(ud.update_workbook, "config", "today", "date", "save_workbook", "sleep")
    result = ud.update_workbook(
        config=_portable_cfg(workbook), today="- #101 First task", date=datetime.date(2026, 7, 27),
        save_workbook=save_once_locked, sleep=lambda seconds: sleeps.append(seconds),
    )

    assert result["status"] == "UPDATED"
    assert len(saves) == 2
    assert sleeps == [result["retry_delay_seconds"]]
    assert openpyxl.load_workbook(workbook)["Daily Report"]["C2"].value == "- #101 First task"


# TC-056: Two locked saves return an actionable failure and leave the original workbook intact.
# Steps:
#   1. Inject two save lock failures.
#   2. Attempt a workbook update.
#   3. Verify structured WORKBOOK_LOCKED and no partial row on disk.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-9, AC-12.
def test_tc_056_reports_two_workbook_lock_failures_without_corrupting_disk_workbook(tmp_path):
    workbook = tmp_path / "DailyTask.xlsx"
    _headers_only_workbook(workbook)
    before = workbook.read_bytes()
    attempts = []
    _require_parameters(ud.update_workbook, "config", "today", "date", "save_workbook", "sleep")

    result = ud.update_workbook(
        config=_portable_cfg(workbook), today="- #101 First task", date=datetime.date(2026, 7, 27),
        save_workbook=lambda _wb, _path: attempts.append(True) or (_ for _ in ()).throw(PermissionError("locked")),
        sleep=lambda _seconds: None,
    )

    assert result["status"] == "FAIL"
    assert result["code"] == "WORKBOOK_LOCKED"
    assert "close" in result["message"].lower()
    assert len(attempts) == 2
    assert workbook.read_bytes() == before


# TC-057: The portable config's Excel path is the sole workbook path used.
# Steps:
#   1. Put a decoy workbook beside the caller and configure a different workbook path.
#   2. Update using the portable configuration.
#   3. Verify only the configured workbook changes.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-5, AC-9, AC-12.
def test_tc_057_uses_workbook_path_from_portable_config_not_caller_directory(tmp_path, monkeypatch):
    configured = tmp_path / "state" / "DailyTask.xlsx"
    configured.parent.mkdir()
    decoy = tmp_path / "DailyTask.xlsx"
    _headers_only_workbook(configured)
    _headers_only_workbook(decoy)
    monkeypatch.chdir(tmp_path)

    ud.update_workbook(
        config=_portable_cfg(configured), today="- #101 First task", date=datetime.date(2026, 7, 27),
        workbook_loader=openpyxl.load_workbook, save_workbook=lambda wb, path: wb.save(path),
    )

    assert openpyxl.load_workbook(configured)["Daily Report"].max_row == 2
    assert openpyxl.load_workbook(decoy)["Daily Report"].max_row == 1


# TC-058: Optional backup is disabled by default and invokes no Git runner.
# Steps:
#   1. Omit backup enablement from portable config.
#   2. Request backup with a runner recorder.
#   3. Verify a structured skip and no argv invocation.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-9, AC-12.
def test_tc_058_skips_backup_by_default_without_running_git(tmp_path):
    workbook = tmp_path / "DailyTask.xlsx"
    workbook.write_bytes(b"workbook")
    commands = []

    _require_parameters(cw.backup_workbook, "config", "runner")
    result = cw.backup_workbook(config=_portable_cfg(workbook), runner=lambda argv: commands.append(argv))

    assert result["status"] == "SKIP"
    assert result["code"] == "BACKUP_DISABLED"
    assert commands == []


# TC-059: Enabled backup outside its configured Git repo is a structured skip, not a failure.
# Steps:
#   1. Enable backup for a temporary workbook outside the injected repo root.
#   2. Run backup through an argv-only runner.
#   3. Verify OUTSIDE_CONFIGURED_REPO and no staging command.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-9, AC-12.
def test_tc_059_skips_enabled_backup_when_workbook_is_outside_configured_repo(tmp_path):
    workbook = tmp_path / "outside" / "DailyTask.xlsx"
    workbook.parent.mkdir()
    workbook.write_bytes(b"workbook")
    commands = []
    cfg = _portable_cfg(workbook)
    cfg["backup"] = {"enabled": True, "repo": str(tmp_path / "repo")}

    result = cw.backup_workbook(config=cfg, runner=lambda argv: commands.append(argv))

    assert result["status"] == "SKIP"
    assert result["code"] == "OUTSIDE_CONFIGURED_REPO"
    assert commands == []


# TC-060: Enabled backup stages and commits only the workbook, preserving unrelated index/worktree state.
# Steps:
#   1. Enable backup and inject Git argv results with unrelated staged/modified paths.
#   2. Back up the configured workbook.
#   3. Verify no broad add, workbook-only paths, and preserved unrelated state.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-9, AC-12.
def test_tc_060_backs_up_only_workbook_without_broad_add_or_unrelated_state_loss(tmp_path):
    repo = tmp_path / "repo"
    repo.mkdir()
    workbook = repo / "DailyTask.xlsx"
    workbook.write_bytes(b"workbook")
    cfg = _portable_cfg(workbook)
    cfg["backup"] = {"enabled": True, "repo": str(repo)}
    commands = []

    def runner(argv):
        commands.append(argv)
        if argv[3:] == ["diff", "--cached", "--quiet", "--", "DailyTask.xlsx"]:
            return {"returncode": 1, "stdout": "", "stderr": ""}
        if argv[3:] == ["rev-parse", "--short", "HEAD"]:
            return {"returncode": 0, "stdout": "abc123\n", "stderr": ""}
        return {"returncode": 0, "stdout": "", "stderr": ""}

    result = cw.backup_workbook(config=cfg, runner=runner)

    assert result == {"status": "COMMITTED", "path": "DailyTask.xlsx", "commit": "abc123"}
    assert ["git", "-C", str(repo), "add", "--", "DailyTask.xlsx"] in commands
    assert not any(command[-1:] == ["."] or command[3:5] == ["add", "."] for command in commands)
    commit = next(command for command in commands if command[3] == "commit")
    assert commit[-2:] == ["--", "DailyTask.xlsx"]


# TC-061: A Git failure is structured and does not escalate to a broad staging recovery.
# Steps:
#   1. Enable backup and inject a failing workbook-only Git add result.
#   2. Run backup.
#   3. Verify a structured failure, actionable error, and no additional Git mutation.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-9, AC-12.
def test_tc_061_reports_backup_git_failure_without_broad_staging_recovery(tmp_path):
    repo = tmp_path / "repo"
    repo.mkdir()
    workbook = repo / "DailyTask.xlsx"
    workbook.write_bytes(b"workbook")
    cfg = _portable_cfg(workbook)
    cfg["backup"] = {"enabled": True, "repo": str(repo)}
    commands = []

    result = cw.backup_workbook(
        config=cfg,
        runner=lambda argv: commands.append(argv) or {"returncode": 1, "stdout": "", "stderr": "index locked"},
    )

    assert result["status"] == "FAIL"
    assert result["code"] == "BACKUP_FAILED"
    assert "index locked" in result["message"]
    assert len(commands) == 1
    assert commands[0][-2:] == ["--", "DailyTask.xlsx"]


# TC-062: Save through a sibling temporary workbook and atomically replace only after save succeeds.
# Steps:
#   1. Inject save and replace boundaries for a configured temporary workbook.
#   2. Update a headers-only workbook, with replace succeeding or failing.
#   3. Verify the save target is a sibling temporary path, replace runs once, and failure preserves final bytes.
# Design: portable-git-daily-report-dev-workflow.md Task 11, AC-9, AC-12.
@pytest.mark.parametrize("replace_fails", [False, True], ids=["replace-success", "replace-failure"])
def test_tc_062_saves_to_sibling_temp_then_replaces_final_without_corruption(tmp_path, replace_fails):
    workbook = tmp_path / "DailyTask.xlsx"
    _headers_only_workbook(workbook)
    original_bytes = workbook.read_bytes()
    save_paths, replace_calls = [], []

    def save_workbook(_workbook, path):
        save_paths.append(Path(path))

    def replace_file(temp_path, final_path):
        replace_calls.append((Path(temp_path), Path(final_path)))
        if replace_fails:
            raise OSError("atomic replacement unavailable")

    _require_parameters(
        ud.update_workbook,
        "config", "today", "date", "save_workbook", "replace_file",
    )
    result = ud.update_workbook(
        config=_portable_cfg(workbook), today="- #101 First task", date=datetime.date(2026, 7, 27),
        save_workbook=save_workbook, replace_file=replace_file,
    )

    assert len(save_paths) == 1
    temporary_path = save_paths[0]
    assert temporary_path.parent == workbook.parent
    assert temporary_path != workbook
    assert len(replace_calls) == 1
    assert replace_calls[0] == (temporary_path, workbook)
    if replace_fails:
        assert result["status"] == "FAIL"
        assert result["code"] == "WORKBOOK_SAVE_FAILED"
        assert "replacement" in result["message"].lower()
        assert workbook.read_bytes() == original_bytes
    else:
        assert result["status"] == "UPDATED"


if __name__ == "__main__":
    tests = sorted(n for n in globals() if n.startswith("test_"))
    failed = 0
    for name in tests:
        try:
            globals()[name]()
            print(f"PASS  {name}")
        except Exception as exc:
            failed += 1
            print(f"FAIL  {name}: {exc}")
    print(f"\n{len(tests) - failed}/{len(tests)} passed")
    sys.exit(1 if failed else 0)


# --- Same-day idempotence -------------------------------------------------------
# A re-run must refresh the existing day's row rather than add a second row for the
# same date, and must not repeat a task line that is already recorded.


def test_repeated_same_day_runs_reuse_one_row_and_do_not_duplicate_lines(tmp_path):
    """Four same-day runs stay on one row; overlapping lines are merged once."""
    workbook_path = tmp_path / "DailyTask.xlsx"
    _headers_only_workbook(workbook_path)
    config = _portable_cfg(workbook_path)
    day = datetime.date(2026, 7, 28)

    modes = [
        ud.update_workbook(config, today, date=day)["mode"]
        for today in ("- #1 Alpha", "- #1 Alpha", "- #1 Alpha\n- #2 Beta", "- #2 Beta")
    ]

    assert modes == ["INSERT", "UPDATE", "UPDATE", "UPDATE"]
    sheet = openpyxl.load_workbook(workbook_path)["Daily Report"]
    assert sheet.max_row == 2, "a same-day re-run must not insert another row"
    lines = str(sheet["C2"].value).splitlines()
    assert lines == ["- #1 Alpha", "- #2 Beta"]


def test_same_day_refresh_keeps_the_date_cell_recognisable(tmp_path):
    """The date cell must not inherit General from a row that does not exist.

    With a single data row there is no row beneath it, and a missing cell reports
    General. Adopting that serialises the date as a bare number, so the next run stops
    recognising the row and inserts a duplicate for the same day.
    """
    workbook_path = tmp_path / "DailyTask.xlsx"
    _headers_only_workbook(workbook_path)
    config = _portable_cfg(workbook_path)
    day = datetime.date(2026, 7, 28)

    ud.update_workbook(config, "- #1 Alpha", date=day)
    ud.update_workbook(config, "- #1 Alpha", date=day)

    sheet = openpyxl.load_workbook(workbook_path)["Daily Report"]
    assert ud._as_date(sheet["A2"].value) == day
    assert sheet["A2"].number_format != "General"


def test_workbook_damaged_by_a_lost_date_format_still_matches_its_row(tmp_path):
    """A date cell already reduced to a bare serial is matched and repaired."""
    workbook_path = tmp_path / "DailyTask.xlsx"
    _headers_only_workbook(workbook_path)
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook["Daily Report"]
    sheet["A2"] = 46231  # 2026-07-28 with the date format lost
    sheet["A2"].number_format = "General"
    sheet["C2"] = "- #1 Alpha"
    workbook.save(workbook_path)

    result = ud.update_workbook(_portable_cfg(workbook_path), "- #2 Beta", date=datetime.date(2026, 7, 28))

    assert result["mode"] == "UPDATE"
    refreshed = openpyxl.load_workbook(workbook_path)["Daily Report"]
    assert refreshed.max_row == 2, "a damaged row must be reused, not duplicated"
    assert str(refreshed["C2"].value).splitlines() == ["- #1 Alpha", "- #2 Beta"]


def test_excel_serial_decoding_round_trips_workbook_dates():
    """Serials decode back to their own date, and non-dates stay unrecognised.

    The 1899-12-30 epoch is exact for any date after 1900-02-28; Excel's phantom
    29 February 1900 shifts only the first two serials, which no report date reaches.
    """
    for expected in (datetime.date(2026, 7, 28), datetime.date(2024, 10, 1), datetime.date(1900, 3, 1)):
        serial = (expected - ud._EXCEL_EPOCH.date()).days
        assert ud._as_date(serial) == expected, f"serial {serial} should decode to {expected}"

    assert ud._as_date(datetime.datetime(2026, 7, 28, 13, 45)) == datetime.date(2026, 7, 28)
    assert ud._as_date(datetime.date(2026, 7, 28)) == datetime.date(2026, 7, 28)
    assert ud._as_date("2026-07-28") is None
    assert ud._as_date(None) is None
    assert ud._as_date(True) is None, "a bool must not be read as a date serial"
