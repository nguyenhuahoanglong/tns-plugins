"""Guardrail for the daily-report skill — verifies the day's outputs against the acceptance criteria.

Checks (deterministic):
  1. Excel row 2 is today's date, with a real datetime + preserved number format.
  2. Yesterday / Today / Report / Timesheet-memo columns are all populated.
  3. The report (column E) matches the "Yesterday\\n...\\nToday\\n..." template.
  4. Unless --skip-timesheet: a timesheet detail line exists for today in the portal.

Exit 0 = all pass; exit 1 = at least one FAIL.

Usage:
  python verify_output.py [--date YYYY-MM-DD] [--config PATH] [--skip-timesheet]
"""
import argparse, datetime, sys, zipfile
import openpyxl
from lib_common import load_config

PASS, FAIL = "PASS", "FAIL"


def _as_date(value):
    return value.date() if isinstance(value, datetime.datetime) else value if isinstance(value, datetime.date) else None


def verify_timesheet_output(cfg, today, expected_description, *, cache_path=None,
                            token_provider=None, dataverse_factory=None, whoami=None):
    """Independently prove one exact detail under this user's one active header."""
    from lib_common import get_access_token, Dataverse, who_am_i
    from write_timesheet import _portable_header, verify_written_detail

    token_provider = token_provider or get_access_token
    dataverse_factory = dataverse_factory or Dataverse
    whoami = whoami or who_am_i
    token = token_provider(cfg, interactive=False, cache_path=cache_path)
    identity = whoami(cfg, token)
    employee_id = identity.get("user_id") if isinstance(identity, dict) else identity
    dv = dataverse_factory(cfg, token)
    headers = _portable_header(dv, cfg, today, employee_id)
    if len(headers) != 1:
        return {"ok": False, "count": 0, "header_count": len(headers)}
    header_id = headers[0].get("id") or headers[0].get("cr90e_xts_timesheet_timesheetheaderid")
    verification = verify_written_detail(dv, cfg, header_id, today, employee_id, expected_description)
    verification["header_count"] = 1
    return verification


def verify_run_result(result, config=None, date=None):
    """Reopen the saved workbook and require post-write Dataverse evidence."""
    checks = []
    workbook = result.get("workbook") if isinstance(result, dict) else None
    workbook_path = (workbook or {}).get("path") if isinstance(workbook, dict) else None
    if not workbook_path and isinstance(workbook, dict):
        workbook_path = workbook.get("workbook")
    checks.append((bool(workbook_path), "portable-path"))
    if workbook_path:
        try:
            reopened = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
            sheet = reopened[(config or {}).get("excel", {}).get("sheet", "Daily Report")]
            expected_date = date or result.get("date")
            if isinstance(expected_date, str):
                expected_date = datetime.datetime.strptime(expected_date, "%Y-%m-%d").date()
            yesterday, today, report_cell = sheet["B2"].value, sheet["C2"].value, sheet["E2"].value
            expected_report = result.get("report", "")
            checks.extend(((_as_date(sheet["A2"].value) == expected_date, "workbook-date"),
                           (yesterday is not None, "workbook-yesterday"),
                           (bool(str(today or "").strip()), "workbook-today"),
                           (report_cell == expected_report and report_cell == f"Yesterday\n{yesterday}\nToday\n{today}", "workbook-report")))
        except (OSError, KeyError, zipfile.BadZipFile):
            checks.append((False, "workbook-reopen"))
    checks.append((bool(workbook) and (workbook.get("updated", True) is not False), "workbook"))
    report = result.get("report", "") if isinstance(result, dict) else ""
    checks.append((isinstance(report, str) and report.startswith("Yesterday\n") and "\nToday\n" in report, "report"))
    timesheet = result.get("timesheet", {}) if isinstance(result, dict) else {}
    queue = result.get("queue", {}) if isinstance(result, dict) else {}
    queued = bool(queue.get("queued"))
    checks.append((bool(timesheet.get("action") or queued), "timesheet-or-queue"))
    if queued:
        record, expected_date = queue.get("record", {}), result.get("date")
        checks.append((isinstance(record, dict) and record.get("date") == str(expected_date) and
                       record.get("todayBlock") == (workbook or {}).get("today") and bool(record.get("lastError") or record.get("reason")),
                       "queue-evidence"))
    if timesheet:
        post_write = timesheet.get("post_write_verification", {})
        checks.append((timesheet.get("status") != "COMMITTED" or bool(post_write.get("ok")), "timesheet-post-write"))
    operations = result.get("operations", []) if isinstance(result, dict) else []
    required = ["gather", "auth", "timesheet_preview", "workbook", "queue" if queued else "timesheet_write", "verify"]
    positions = {name: operations.index(name) for name in required if name in operations}
    ordered = all(name in positions for name in required) and all(
        positions[left] < positions[right] for left, right in zip(required, required[1:]))
    checks.append((ordered, "operation-order"))
    return {"ok": all(ok for ok, _ in checks), "checks": [name for ok, name in checks if ok],
            "failures": [name for ok, name in checks if not ok]}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date")
    ap.add_argument("--config")
    ap.add_argument("--skip-timesheet", action="store_true")
    args = ap.parse_args()

    cfg = load_config(args.config)
    today = (datetime.datetime.strptime(args.date, "%Y-%m-%d").date()
             if args.date else datetime.date.today())
    results = []

    # --- Excel checks ---
    wb = openpyxl.load_workbook(cfg["excel"]["path"])
    ws = wb[cfg["excel"]["sheet"]]
    a2, b2, c2, e2, f2 = (ws["A2"].value, ws["B2"].value, ws["C2"].value,
                          ws["E2"].value, ws["F2"].value)

    a2_date = a2.date() if isinstance(a2, datetime.datetime) else None
    results.append((a2_date == today, "Excel row 2 date is today",
                    f"got {a2_date!r}, expected {today}"))
    results.append((isinstance(a2, datetime.datetime) and ws["A2"].number_format not in (None, "General"),
                    "Date cell keeps a date number format", f"format={ws['A2'].number_format!r}"))
    results.append((bool(b2 and str(b2).strip()), "Yesterday populated", repr(b2)[:60]))
    results.append((bool(c2 and str(c2).strip()), "Today populated", repr(c2)[:60]))
    results.append((bool(f2 and "Task Description:" in str(f2)), "Timesheet memo populated", repr(f2)[:60]))

    expected_report = f"Yesterday\n{b2}\nToday\n{c2}"
    results.append((e2 == expected_report, "Report matches Yesterday/Today template",
                    "column E differs from composed report"))

    # --- Timesheet check ---
    if not args.skip_timesheet:
        try:
            from lib_common import resolve_runtime_context
            from write_timesheet import format_description
            context = resolve_runtime_context(config_path=args.config)
            ts = cfg["timesheet"]
            expected_description = format_description(str(c2 or ""), ts["defaults"].get("description_style", "semicolon"))
            verification = verify_timesheet_output(cfg, today, expected_description, cache_path=context.auth_cache_path)
            exact = verification["ok"]
            results.append((exact, "Timesheet has exactly one matching line with exact description",
                            f"found {verification['count']} line(s), expected description={expected_description!r}"))
        except SystemExit as e:
            results.append((False, "Timesheet has exactly one matching line with exact description", f"auth/query unavailable: {e}"))

    # --- report ---
    ok = all(r[0] for r in results)
    print("=== daily-report verify ===")
    for passed, name, detail in results:
        print(f"  [{PASS if passed else FAIL}] {name}" + ("" if passed else f"  ({detail})"))
    print(f"=== {'ALL PASS' if ok else 'FAILURES PRESENT'} ===")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
